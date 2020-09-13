import json
from pathlib import Path
from random import *

from openpyxl import load_workbook
import codecs


def ReadAndConvert():
    '''
        read disease data from the assets/disease.xlsx
    '''

    Disease_Data = {}

    # convert from xlsx to dict
    file_path = 'assets/disease.xlsx'
    wb = load_workbook(file_path)

    for plantName in wb.sheetnames:
        sheet: Worksheet = wb[plantName]
        Disease_Data[plantName] = []

        for diseaseRange in sheet.merged_cells.ranges:
            Disease_Data[plantName].append(
                buildDiseaseFromRange(
                    sheet,
                    range(diseaseRange.min_row, diseaseRange.max_row)
                )
            )


    # save data 
    file = codecs.open('db/disease_data.py', 'wb', 'utf-8')
    JSON_Disease_Data = json.dumps(Disease_Data, ensure_ascii=False, indent=3)
    file.write('Disease_Data = ')
    file.write(JSON_Disease_Data)

    plants_Array = wb.sheetnames
    JSON_plants_Array = json.dumps(plants_Array, ensure_ascii=False, indent=3)
    file.write('\nplants_Array = ')
    file.write(JSON_plants_Array)
    file.close()

    return plants_Array, Disease_Data


def buildDiseaseFromRange(sheet, range):
    '''
        get one disease data after getting its range its range
    '''

    result_disease = {
        'name': sheet.cell(row=range[0], column=1).value,
        'link': sheet.cell(row=range[-1], column=2).value,
    }

    # build details
    details = []
    for row in range[0:-1]:
        detail = sheet.cell(row=row, column=2).value
        if detail != None:
            details.append(detail)

    result_disease['details'] = '\n'.join(details)

    symptoms = []
    # build sumptoms
    for row in range:
        symptomNameAr = sheet.cell(row=row, column=3).value
        symptomName = sheet.cell(row=row, column=4).value
        symptomCF = sheet.cell(row=row, column=5).value
        if symptomName != None:
            symptoms.append({
                'name': symptomName,
                'nameAr': symptomNameAr,
                # generate random cf in case there isn't
                'CF': float(symptomCF) if symptomCF != None else random()
            })

    result_disease['symptoms'] = symptoms
    # print(result_disease)
    return result_disease

